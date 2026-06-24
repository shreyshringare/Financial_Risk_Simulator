"""
Excel report exporter for financial risk analysis results.
Produces a formatted multi-sheet workbook with charts.
"""
import os
from datetime import datetime
from typing import Optional
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ACCENT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DANGER_FILL = PatternFill(start_color="FFDDC1", end_color="FFDDC1", fill_type="solid")  # for losses
BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _style_header_row(ws, row: int, n_cols: int) -> None:
    """Apply HEADER_FILL, HEADER_FONT, center alignment and BORDER to a row."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _auto_width(ws) -> None:
    """Set column widths based on max content length, capped at 40."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        adjusted = min(max_len + 2, 40)
        ws.column_dimensions[col_letter].width = adjusted


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------

def export_risk_report(
    ticker: str,
    simulated_paths: np.ndarray,
    risk_metrics: dict,
    stress_results: Optional[pd.DataFrame] = None,
    correlation_df: Optional[pd.DataFrame] = None,
    output_dir: str = "./reports"
) -> str:
    """
    Generate a formatted Excel risk report.

    Args:
        ticker: Stock ticker symbol
        simulated_paths: shape (simulations, days) Monte Carlo paths
        risk_metrics: dict with var, cvar, sharpe, max_drawdown
        stress_results: DataFrame from compare_scenarios() (optional)
        correlation_df: Correlation matrix DataFrame (optional)
        output_dir: Directory to save report

    Returns:
        Full path to the saved .xlsx file.
    """
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ------------------------------------------------------------------
    # Sheet 1: Summary
    # ------------------------------------------------------------------
    ws_summary = wb.create_sheet("Summary")

    # Title row (merged A1:E1)
    ws_summary.merge_cells("A1:E1")
    title_cell = ws_summary["A1"]
    title_cell.value = f"Financial Risk Report — {ticker}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 30

    # Generated timestamp
    ws_summary["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_summary["A2"].font = Font(italic=True, size=10)

    # Row 3 blank
    ws_summary["A3"] = ""

    # Headers row 4
    headers = ["Metric", "Value", "Interpretation"]
    for col_idx, h in enumerate(headers, start=1):
        ws_summary.cell(row=4, column=col_idx).value = h
    _style_header_row(ws_summary, row=4, n_cols=len(headers))

    var_val = risk_metrics.get("var", 0)
    cvar_val = risk_metrics.get("cvar", 0)
    sharpe_val = risk_metrics.get("sharpe", 0)
    max_dd_val = risk_metrics.get("max_drawdown", 0)

    metrics_rows = [
        ("VaR (95%)", f"{var_val:.2%}", "95% of days, loss won't exceed this"),
        ("CVaR (95%)", f"{cvar_val:.2%}", "Average loss in worst 5% scenarios"),
        ("Sharpe Ratio", f"{sharpe_val:.4f}", ">1.0 good, >2.0 excellent, <0 poor"),
        ("Max Drawdown", f"{max_dd_val:.2%}", "Largest peak-to-trough decline"),
    ]

    for row_offset, (metric, value, interp) in enumerate(metrics_rows):
        row_num = 5 + row_offset
        ws_summary.cell(row=row_num, column=1).value = metric
        val_cell = ws_summary.cell(row=row_num, column=2)
        val_cell.value = value
        ws_summary.cell(row=row_num, column=3).value = interp

        # Style all cells in row
        for col_idx in range(1, 4):
            c = ws_summary.cell(row=row_num, column=col_idx)
            c.border = BORDER
            c.alignment = Alignment(horizontal="left", vertical="center")

        # DANGER_FILL on VaR and CVaR value cells
        if metric in ("VaR (95%)", "CVaR (95%)"):
            val_cell.fill = DANGER_FILL

    # Disclaimer note at row 10
    ws_summary["A10"] = (
        "⚠ VaR/CVaR assume GBM (log-normal returns, constant volatility). "
        "Fat tails and volatility clustering are not captured."
    )
    ws_summary["A10"].font = Font(italic=True, size=9, color="7F7F7F")
    ws_summary.merge_cells("A10:E10")

    _auto_width(ws_summary)

    # ------------------------------------------------------------------
    # Sheet 2: Monte Carlo Paths
    # ------------------------------------------------------------------
    ws_mc = wb.create_sheet("Monte Carlo Paths")

    mc_headers = [
        "Simulation #", "Day 1 Price", "Day 63 Price",
        "Day 126 Price", "Day 189 Price", "Final Price", "Return %"
    ]
    for col_idx, h in enumerate(mc_headers, start=1):
        ws_mc.cell(row=1, column=col_idx).value = h
    _style_header_row(ws_mc, row=1, n_cols=len(mc_headers))

    n_sims, n_days = simulated_paths.shape
    display_sims = min(100, n_sims)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for i in range(display_sims):
        row_num = i + 2
        day1 = simulated_paths[i, 0]
        day63 = simulated_paths[i, 62] if n_days > 62 else simulated_paths[i, -1]
        day126 = simulated_paths[i, 125] if n_days > 125 else simulated_paths[i, -1]
        day189 = simulated_paths[i, 188] if n_days > 188 else simulated_paths[i, -1]
        final = simulated_paths[i, -1]
        ret_pct = (final / day1 - 1) if day1 != 0 else 0

        row_data = [i + 1, round(day1, 4), round(day63, 4),
                    round(day126, 4), round(day189, 4), round(final, 4),
                    f"{ret_pct:.2%}"]

        for col_idx, val in enumerate(row_data, start=1):
            c = ws_mc.cell(row=row_num, column=col_idx)
            c.value = val
            c.border = BORDER
            c.alignment = Alignment(horizontal="right")

        # Color Return % cell
        ret_cell = ws_mc.cell(row=row_num, column=7)
        ret_cell.fill = green_fill if ret_pct >= 0 else DANGER_FILL

    # LineChart — first 20 paths
    chart_paths = min(20, n_sims)
    # Write price data for chart on a hidden area starting at column 9
    ws_mc.cell(row=1, column=9).value = "ChartDay"
    for d in range(n_days):
        ws_mc.cell(row=d + 2, column=9).value = d + 1

    for i in range(chart_paths):
        ws_mc.cell(row=1, column=10 + i).value = f"Sim {i + 1}"
        for d in range(n_days):
            ws_mc.cell(row=d + 2, column=10 + i).value = round(float(simulated_paths[i, d]), 4)

    lc = LineChart()
    lc.title = f"Monte Carlo Paths — {ticker} (first {chart_paths} simulations)"
    lc.style = 10
    lc.y_axis.title = "Price"
    lc.x_axis.title = "Day"
    lc.width = 30
    lc.height = 18

    for i in range(chart_paths):
        data_ref = Reference(ws_mc, min_col=10 + i, min_row=1, max_row=n_days + 1)
        series = openpyxl.chart.Series(data_ref, title=f"Sim {i + 1}")
        lc.series.append(series)

    cats = Reference(ws_mc, min_col=9, min_row=2, max_row=n_days + 1)
    lc.set_categories(cats)
    ws_mc.add_chart(lc, "A103")

    _auto_width(ws_mc)

    # ------------------------------------------------------------------
    # Sheet 3: Return Distribution
    # ------------------------------------------------------------------
    ws_dist = wb.create_sheet("Return Distribution")

    terminal_returns = (simulated_paths[:, -1] / simulated_paths[:, 0]) - 1
    n_bins = 20
    counts, bin_edges = np.histogram(terminal_returns, bins=n_bins)
    total_sims = len(terminal_returns)

    dist_headers = ["Bucket (Return Range)", "Count", "% of Simulations"]
    for col_idx, h in enumerate(dist_headers, start=1):
        ws_dist.cell(row=1, column=col_idx).value = h
    _style_header_row(ws_dist, row=1, n_cols=len(dist_headers))

    for i in range(n_bins):
        row_num = i + 2
        bucket_label = f"{bin_edges[i]:.2%} to {bin_edges[i+1]:.2%}"
        count = int(counts[i])
        pct = count / total_sims if total_sims > 0 else 0

        ws_dist.cell(row=row_num, column=1).value = bucket_label
        ws_dist.cell(row=row_num, column=2).value = count
        ws_dist.cell(row=row_num, column=3).value = f"{pct:.2%}"

        for col_idx in range(1, 4):
            c = ws_dist.cell(row=row_num, column=col_idx)
            c.border = BORDER
            c.alignment = Alignment(horizontal="right")

    # BarChart for distribution
    bc = BarChart()
    bc.type = "col"
    bc.title = f"Return Distribution — {ticker}"
    bc.y_axis.title = "Count"
    bc.x_axis.title = "Return Bucket"
    bc.width = 25
    bc.height = 15

    data_ref = Reference(ws_dist, min_col=2, min_row=1, max_row=n_bins + 1)
    bc.add_data(data_ref, titles_from_data=True)
    cats_ref = Reference(ws_dist, min_col=1, min_row=2, max_row=n_bins + 1)
    bc.set_categories(cats_ref)
    ws_dist.add_chart(bc, "E2")

    # VaR threshold note
    var_note_row = n_bins + 4
    ws_dist.cell(row=var_note_row, column=1).value = (
        f"VaR threshold (95%): {var_val:.2%}  — "
        "Returns to the left of this value represent the worst-case 5% tail."
    )
    ws_dist.cell(row=var_note_row, column=1).font = Font(italic=True, size=9, color="7F7F7F")
    ws_dist.merge_cells(
        start_row=var_note_row, start_column=1,
        end_row=var_note_row, end_column=3
    )

    _auto_width(ws_dist)

    # ------------------------------------------------------------------
    # Sheet 4: Stress Tests (optional)
    # ------------------------------------------------------------------
    if stress_results is not None:
        ws_stress = wb.create_sheet("Stress Tests")

        stress_headers = [
            "Scenario", "Description", "Shock",
            "Baseline VaR", "Stressed VaR", "Worst Case", "% Paths >50% Loss"
        ]
        for col_idx, h in enumerate(stress_headers, start=1):
            ws_stress.cell(row=1, column=col_idx).value = h
        _style_header_row(ws_stress, row=1, n_cols=len(stress_headers))

        col_map = {
            "scenario": 1, "description": 2, "shock": 3,
            "baseline_var_95": 4, "stressed_var_95": 5,
            "worst_case": 6, "pct_paths_over_50pct_loss": 7
        }

        for df_row_idx, (_, df_row) in enumerate(stress_results.iterrows()):
            row_num = df_row_idx + 2
            for col_name, col_idx in col_map.items():
                val = df_row.get(col_name, "")
                c = ws_stress.cell(row=row_num, column=col_idx)
                c.value = val
                c.border = BORDER
                c.alignment = Alignment(horizontal="right")

            # DANGER_FILL for severe stress rows
            stressed_var = df_row.get("stressed_var_95", 0)
            if pd.notna(stressed_var) and float(stressed_var) < -0.4:
                for col_idx in range(1, len(stress_headers) + 1):
                    ws_stress.cell(row=row_num, column=col_idx).fill = DANGER_FILL

        _auto_width(ws_stress)

    # ------------------------------------------------------------------
    # Sheet 5: Correlation (optional)
    # ------------------------------------------------------------------
    if correlation_df is not None:
        ws_corr = wb.create_sheet("Correlation")

        tickers = list(correlation_df.columns)
        n_tickers = len(tickers)

        # Column headers (row 1, starting col 2)
        for col_idx, t in enumerate(tickers, start=2):
            c = ws_corr.cell(row=1, column=col_idx)
            c.value = t
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center")
            c.border = BORDER

        # Row headers (col 1, starting row 2) + data
        high_pos_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        neg_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row_offset, row_ticker in enumerate(tickers):
            row_num = row_offset + 2
            # Row label
            label_cell = ws_corr.cell(row=row_num, column=1)
            label_cell.value = row_ticker
            label_cell.fill = HEADER_FILL
            label_cell.font = HEADER_FONT
            label_cell.alignment = Alignment(horizontal="center")
            label_cell.border = BORDER

            for col_offset, col_ticker in enumerate(tickers):
                col_num = col_offset + 2
                corr_val = correlation_df.loc[row_ticker, col_ticker]
                c = ws_corr.cell(row=row_num, column=col_num)
                c.value = round(float(corr_val), 4)
                c.border = BORDER
                c.alignment = Alignment(horizontal="center")

                # Color coding
                if corr_val > 0.7:
                    c.fill = high_pos_fill
                elif corr_val < -0.3:
                    c.fill = neg_fill
                # else: white / default

        _auto_width(ws_corr)

    # ------------------------------------------------------------------
    # Save workbook
    # ------------------------------------------------------------------
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{ticker}_risk_report_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)

    return filepath
