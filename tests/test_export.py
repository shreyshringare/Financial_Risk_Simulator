"""
Tests for export/excel_exporter.py and export/powerbi_exporter.py.
Uses real file I/O via pytest's tmp_path fixture.
"""
import numpy as np
import pandas as pd
import pytest
import openpyxl


# ---------------------------------------------------------------------------
# Shared fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def sim_paths():
    """Small deterministic Monte Carlo paths array (50 sims × 10 days)."""
    np.random.seed(0)
    start = 100.0
    returns = np.random.normal(0.0005, 0.01, size=(50, 10))
    paths = start * np.exp(np.cumsum(returns, axis=1))
    # Prepend start column so path[i, 0] == 100
    paths = np.hstack([np.full((50, 1), start), paths])
    return paths


@pytest.fixture()
def risk_metrics():
    return {"var": -0.02, "cvar": -0.035, "sharpe": 1.2, "max_drawdown": -0.15}


@pytest.fixture()
def prices():
    idx = pd.date_range("2023-01-02", periods=50, freq="B")
    np.random.seed(1)
    return pd.Series(100 + np.cumsum(np.random.randn(50)), index=idx, name="Close")


@pytest.fixture()
def stress_df():
    return pd.DataFrame([{
        "scenario": "2008_financial_crisis",
        "description": "GFC",
        "shock": -0.55,
        "baseline_var_95": -0.02,
        "stressed_var_95": -0.45,
        "worst_case": -0.80,
        "pct_paths_over_50pct_loss": 0.12,
    }])


@pytest.fixture()
def correlation_df():
    return pd.DataFrame(
        [[1.0, 0.8], [0.8, 1.0]],
        index=["AAPL", "MSFT"],
        columns=["AAPL", "MSFT"],
    )


# ---------------------------------------------------------------------------
# excel_exporter tests
# ---------------------------------------------------------------------------

class TestExcelExporter:

    def test_creates_xlsx_file(self, tmp_path, sim_paths, risk_metrics):
        from export.excel_exporter import export_risk_report

        path = export_risk_report("AAPL", sim_paths, risk_metrics, output_dir=str(tmp_path))

        assert path.endswith(".xlsx")
        assert (tmp_path / path.split("\\")[-1].split("/")[-1]).exists()

    def test_summary_sheet_exists(self, tmp_path, sim_paths, risk_metrics):
        from export.excel_exporter import export_risk_report

        path = export_risk_report("AAPL", sim_paths, risk_metrics, output_dir=str(tmp_path))
        wb = openpyxl.load_workbook(path)

        assert "Summary" in wb.sheetnames

    def test_summary_title_contains_ticker(self, tmp_path, sim_paths, risk_metrics):
        from export.excel_exporter import export_risk_report

        path = export_risk_report("TSLA", sim_paths, risk_metrics, output_dir=str(tmp_path))
        wb = openpyxl.load_workbook(path)
        ws = wb["Summary"]

        assert "TSLA" in str(ws["A1"].value)

    def test_summary_has_all_four_metrics(self, tmp_path, sim_paths, risk_metrics):
        from export.excel_exporter import export_risk_report

        path = export_risk_report("AAPL", sim_paths, risk_metrics, output_dir=str(tmp_path))
        wb = openpyxl.load_workbook(path)
        ws = wb["Summary"]

        all_values = [str(ws.cell(row=r, column=1).value) for r in range(1, 20)]
        assert any("VaR" in v for v in all_values)
        assert any("CVaR" in v for v in all_values)
        assert any("Sharpe" in v for v in all_values)
        assert any("Drawdown" in v or "drawdown" in v for v in all_values)

    def test_monte_carlo_sheet_exists(self, tmp_path, sim_paths, risk_metrics):
        from export.excel_exporter import export_risk_report

        path = export_risk_report("AAPL", sim_paths, risk_metrics, output_dir=str(tmp_path))
        wb = openpyxl.load_workbook(path)

        assert "Monte Carlo Paths" in wb.sheetnames

    def test_monte_carlo_has_simulation_rows(self, tmp_path, sim_paths, risk_metrics):
        from export.excel_exporter import export_risk_report

        path = export_risk_report("AAPL", sim_paths, risk_metrics, output_dir=str(tmp_path))
        wb = openpyxl.load_workbook(path)
        ws = wb["Monte Carlo Paths"]

        # Row 1 = headers, row 2+ = data; 50 sims → rows 2–51
        assert ws.cell(row=2, column=1).value == 1  # Simulation #1
        assert ws.cell(row=51, column=1).value == 50  # Simulation #50

    def test_return_distribution_sheet_exists(self, tmp_path, sim_paths, risk_metrics):
        from export.excel_exporter import export_risk_report

        path = export_risk_report("AAPL", sim_paths, risk_metrics, output_dir=str(tmp_path))
        wb = openpyxl.load_workbook(path)

        assert "Return Distribution" in wb.sheetnames

    def test_stress_sheet_present_when_provided(self, tmp_path, sim_paths, risk_metrics, stress_df):
        from export.excel_exporter import export_risk_report

        path = export_risk_report("AAPL", sim_paths, risk_metrics, stress_results=stress_df, output_dir=str(tmp_path))
        wb = openpyxl.load_workbook(path)

        assert "Stress Tests" in wb.sheetnames

    def test_stress_sheet_absent_when_not_provided(self, tmp_path, sim_paths, risk_metrics):
        from export.excel_exporter import export_risk_report

        path = export_risk_report("AAPL", sim_paths, risk_metrics, output_dir=str(tmp_path))
        wb = openpyxl.load_workbook(path)

        assert "Stress Tests" not in wb.sheetnames

    def test_correlation_sheet_present_when_provided(self, tmp_path, sim_paths, risk_metrics, correlation_df):
        from export.excel_exporter import export_risk_report

        path = export_risk_report("AAPL", sim_paths, risk_metrics, correlation_df=correlation_df, output_dir=str(tmp_path))
        wb = openpyxl.load_workbook(path)

        assert "Correlation" in wb.sheetnames

    def test_formula_injection_guard_in_correlation(self, tmp_path, sim_paths, risk_metrics):
        """Ticker starting with = must not appear as a formula in the cell."""
        from export.excel_exporter import export_risk_report

        evil_ticker = "=EVIL"
        corr = pd.DataFrame(
            [[1.0, 0.5], [0.5, 1.0]],
            index=[evil_ticker, "MSFT"],
            columns=[evil_ticker, "MSFT"],
        )
        path = export_risk_report("AAPL", sim_paths, risk_metrics, correlation_df=corr, output_dir=str(tmp_path))
        wb = openpyxl.load_workbook(path)
        ws = wb["Correlation"]

        # All cell values in the correlation sheet should NOT start with =
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    assert not str(cell.value).startswith("="), f"Formula injection at {cell.coordinate}: {cell.value}"


# ---------------------------------------------------------------------------
# powerbi_exporter tests
# ---------------------------------------------------------------------------

class TestPowerBIExporter:

    def test_returns_dict_with_expected_keys(self, tmp_path, prices, sim_paths, risk_metrics):
        from export.powerbi_exporter import export_for_powerbi

        result = export_for_powerbi("AAPL", prices, sim_paths, risk_metrics, output_dir=str(tmp_path))

        assert "prices" in result
        assert "risk_metrics" in result
        assert "monte_carlo_summary" in result
        assert "monte_carlo_percentiles" in result
        assert "schema" in result

    def test_all_files_exist_on_disk(self, tmp_path, prices, sim_paths, risk_metrics):
        from export.powerbi_exporter import export_for_powerbi
        import pathlib

        result = export_for_powerbi("AAPL", prices, sim_paths, risk_metrics, output_dir=str(tmp_path))

        for key, path in result.items():
            assert pathlib.Path(path).exists(), f"Missing file for key '{key}': {path}"

    def test_prices_csv_has_required_columns(self, tmp_path, prices, sim_paths, risk_metrics):
        from export.powerbi_exporter import export_for_powerbi

        result = export_for_powerbi("AAPL", prices, sim_paths, risk_metrics, output_dir=str(tmp_path))
        df = pd.read_csv(result["prices"])

        assert "date" in df.columns
        assert "ticker" in df.columns
        assert "close_price" in df.columns
        assert "log_return" in df.columns
        assert (df["ticker"] == "AAPL").all()

    def test_risk_metrics_csv_has_four_rows(self, tmp_path, prices, sim_paths, risk_metrics):
        from export.powerbi_exporter import export_for_powerbi

        result = export_for_powerbi("AAPL", prices, sim_paths, risk_metrics, output_dir=str(tmp_path))
        df = pd.read_csv(result["risk_metrics"])

        assert len(df) == 4
        metric_names = set(df["metric_name"])
        assert "VaR_95" in metric_names
        assert "CVaR_95" in metric_names
        assert "Sharpe_Ratio" in metric_names
        assert "Max_Drawdown" in metric_names

    def test_monte_carlo_summary_row_count(self, tmp_path, prices, sim_paths, risk_metrics):
        from export.powerbi_exporter import export_for_powerbi

        result = export_for_powerbi("AAPL", prices, sim_paths, risk_metrics, output_dir=str(tmp_path))
        df = pd.read_csv(result["monte_carlo_summary"])

        # One row per simulation
        assert len(df) == sim_paths.shape[0]
        assert "simulation_id" in df.columns
        assert "terminal_return_pct" in df.columns
        assert "above_initial" in df.columns

    def test_stress_csv_created_when_provided(self, tmp_path, prices, sim_paths, risk_metrics, stress_df):
        from export.powerbi_exporter import export_for_powerbi

        result = export_for_powerbi("AAPL", prices, sim_paths, risk_metrics, stress_results=stress_df, output_dir=str(tmp_path))

        assert "stress_tests" in result
        df = pd.read_csv(result["stress_tests"])
        assert "ticker" in df.columns
        assert "report_date" in df.columns

    def test_stress_csv_absent_when_not_provided(self, tmp_path, prices, sim_paths, risk_metrics):
        from export.powerbi_exporter import export_for_powerbi

        result = export_for_powerbi("AAPL", prices, sim_paths, risk_metrics, output_dir=str(tmp_path))

        assert "stress_tests" not in result

    def test_correlation_csv_melted_to_long_format(self, tmp_path, prices, sim_paths, risk_metrics, correlation_df):
        from export.powerbi_exporter import export_for_powerbi

        result = export_for_powerbi("AAPL", prices, sim_paths, risk_metrics, correlation_df=correlation_df, output_dir=str(tmp_path))
        df = pd.read_csv(result["correlation"])

        assert "ticker_a" in df.columns
        assert "ticker_b" in df.columns
        assert "correlation" in df.columns
        assert "strength" in df.columns
        # 2×2 matrix → 4 rows
        assert len(df) == 4

    def test_schema_md_created(self, tmp_path, prices, sim_paths, risk_metrics):
        from export.powerbi_exporter import export_for_powerbi
        import pathlib

        result = export_for_powerbi("AAPL", prices, sim_paths, risk_metrics, output_dir=str(tmp_path))
        schema = pathlib.Path(result["schema"]).read_text(encoding="utf-8")

        assert "PowerBI" in schema
        assert "AAPL" in schema
